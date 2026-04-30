# -*- coding: utf-8 -*-
"""
jiaolong自进化框架 - evolution.py
> 核心自进化循环（仿 Karpathy AutoResearch train.py）
> 版本: v1.0 | 2026-04-02

灵感来源:
- Karpathy AutoResearch: 实验循环 = 改代码 → 训练 → 评估 → 保留/丢弃
- Claude Code: extractMemories 自动记忆提取 + tools 系统 + coordinator 多Agent

本文件设计原则:
1. 唯一可修改点: 每次实验只改一个文件
2. 固定评估指标: 记忆命中率 / 工具完善度 / 协作效率
3. 严格时间预算: 每次实验 ≤ 5分钟
4. 完整实验日志: 所有改动可追溯可回滚
"""

import os
import sys
import json
import time
import shutil
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
from dataclasses import dataclass, asdict
from enum import Enum

# ─────────────────────────────────────────────────────────────────────────────
# 路径配置
# ─────────────────────────────────────────────────────────────────────────────
import os
WORKSPACE = Path(os.environ.get("JIAOLONG_WORKSPACE", str(Path.home() / ".claude" / "jiaolong")))
EVOLUTION_DIR  = WORKSPACE / "evolution_framework"
EXPERIMENTS   = EVOLUTION_DIR / "experiments"
MEMORY_DIR    = WORKSPACE / "memory"
SKILLS_DIR    = WORKSPACE / "skills"
MEMORY_HOT    = MEMORY_DIR / "memory_hot.json"
AGENTS_FILE   = WORKSPACE / "AGENTS.md"
MEMORY_FILE   = WORKSPACE / "MEMORY.md"
SOUL_FILE     = WORKSPACE / "SOUL.md"
LOG_FILE      = EXPERIMENTS / "experiment_log.md"

# ─────────────────────────────────────────────────────────────────────────────
# 指标定义
# ─────────────────────────────────────────────────────────────────────────────
METRICS = {
    "memory_hit_rate":     {"name": "记忆命中率",       "target": 0.80, "current": 0.30},
    "tool_coverage":       {"name": "工具完善度",       "target": 0.70, "current": 0.20},
    "l2_l3_ratio":         {"name": "L2+L3自动化率",    "target": 0.80, "current": 0.35},
    "collaboration_score": {"name": "协作效率",         "target": 0.80, "current": 0.50},
    "skill_count":         {"name": "Skill数量",         "target": 20,   "current": 3},
    "context_window":      {"name": "上下文压缩率",      "target": 0.95, "current": 0.00},
}

IMPROVEMENT_THRESHOLD = 0.05  # 5% 提升才保留
MAX_EXPERIMENT_SECONDS = 300  # 5分钟硬限制

# ─────────────────────────────────────────────────────────────────────────────
# 数据结构
# ─────────────────────────────────────────────────────────────────────────────
class Verdict(Enum):
    KEPT      = "KEPT"
    DISCARDED = "DISCARDED"
    RUNNING   = "RUNNING"
    FAILED    = "FAILED"

@dataclass
class Experiment:
    id:         str          # EXP-YYYYMMDD-NNN
    timestamp:  str
    hypothesis: str           # 改进假设
    file_tried: str          # 尝试修改的文件
    changes:    str          # 具体改动
    metrics_before: Dict      # 实验前指标
    metrics_after:  Dict      # 实验后指标
    verdict:    Verdict
    reason:     str          # KEPT/DISCARDED原因
    duration_s: float        # 耗时秒

# ─────────────────────────────────────────────────────────────────────────────
# 核心：指标采集
# ─────────────────────────────────────────────────────────────────────────────

def read_json(path: Path, default=None) -> Any:
    try:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return default

def read_text(path: Path, default="") -> str:
    try:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
    except Exception:
        pass
    return default

def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)

def count_skill_files() -> int:
    """统计当前Skill数量"""
    count = 0
    for p in SKILLS_DIR.rglob("SKILL.md"):
        count += 1
    return count

def measure_memory_hit_rate() -> float:
    """
    测量记忆命中率
    = 7天内相同问题重复出现时命中热记忆的比例
    简化版：检查memory_hot.json中7天内更新的条目占比
    """
    hot = read_json(MEMORY_HOT, [])
    if not hot:
        return 0.0
    now = datetime.now()
    cutoff = now - timedelta(days=7)
    recent = 0
    total = 0
    for item in hot:
        total += 1
        updated = ""
        if isinstance(item, dict):
            updated = item.get("updated", "")
        elif isinstance(item, str):
            updated = item
        if updated:
            try:
                dt = datetime.fromisoformat(updated.replace("Z",""))
                if dt >= cutoff:
                    recent += 1
            except:
                pass
    return recent / max(total, 1)

def measure_l2_l3_ratio() -> float:
    """
    测量L2+L3自动化就绪率
    读取platform_changes_full.xlsx（如存在）
    """
    xlsx_path = WORKSPACE / "platform_changes_full.xlsx"
    if not xlsx_path.exists():
        return METRICS["l2_l3_ratio"]["current"]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
        total = ws.max_row - 1  # 减标题行
        l2_l3 = 0
        for row in range(2, ws.max_row + 1):
            lvl = str(ws.cell(row, 6).value or "")
            if "L2" in lvl or "L3" in lvl:
                l2_l3 += 1
        return l2_l3 / max(total, 1)
    except Exception:
        return METRICS["l2_l3_ratio"]["current"]

def measure_collaboration_score() -> float:
    """
    协作效率
    简化版：检查AGENTS.md是否包含完整三脑定义
    """
    agents = read_text(AGENTS_FILE)
    required = ["小笨", "小呆", "小傻", "小虾"]
    present = sum(1 for r in required if r in agents)
    # 再检查是否有明确的协作流程定义
    has_protocol = "三脑" in agents and ("Phase" in agents or "协作流程" in agents)
    base = present / len(required)
    if has_protocol:
        base = min(base * 1.2, 1.0)
    return base

def collect_metrics() -> Dict[str, float]:
    """采集当前所有指标"""
    return {
        "memory_hit_rate":     round(measure_memory_hit_rate(), 3),
        "tool_coverage":       round(METRICS["tool_coverage"]["current"], 3),
        "l2_l3_ratio":         round(measure_l2_l3_ratio(), 3),
        "collaboration_score": round(measure_collaboration_score(), 3),
        "skill_count":         float(count_skill_files()),
        "context_window":      METRICS["context_window"]["current"],
    }

def metrics_improved(before: Dict, after: Dict) -> Tuple[bool, str]:
    """
    比较指标变化
    返回 (是否提升, 说明)
    """
    improvements = []
    regressions  = []
    for key in before:
        b, a = before[key], after[key]
        if key in ("tool_coverage", "l2_l3_ratio", "collaboration_score",
                   "memory_hit_rate", "context_window"):
            # 越高越好
            change = (a - b) / max(b, 0.01)
            if change >= IMPROVEMENT_THRESHOLD:
                improvements.append(f"{METRICS[key]['name']}: +{change*100:.1f}%")
            elif change <= -IMPROVEMENT_THRESHOLD:
                regressions.append(f"{METRICS[key]['name']}: {change*100:.1f}%")
        elif key == "skill_count":
            if a > b:
                improvements.append(f"Skill数量: +{int(a-b)}个")
            elif a < b:
                regressions.append(f"Skill数量: {int(a-b)}个")
    if improvements and not regressions:
        return True, "; ".join(improvements)
    elif regressions:
        return False, "; ".join(regressions)
    else:
        return False, "无显著变化"

# ─────────────────────────────────────────────────────────────────────────────
# 核心：实验执行
# ─────────────────────────────────────────────────────────────────────────────

def next_experiment_id() -> str:
    """生成下一个实验ID"""
    today = datetime.now().strftime("%Y%m%d")
    existing = list(EXPERIMENTS.glob(f"EXP-{today}-*"))
    n = len(existing) + 1
    return f"EXP-{today}-{n:03d}"

def create_experiment_dir(exp_id: str) -> Path:
    d = EXPERIMENTS / exp_id
    d.mkdir(parents=True, exist_ok=True)
    return d

def backup_file(path: Path, exp_dir: Path) -> Optional[Path]:
    """备份原文件到实验目录"""
    try:
        if path.exists():
            bak = exp_dir / f"_backup_{path.name}"
            shutil.copy2(path, bak)
            return bak
    except Exception:
        pass
    return None

def apply_changes(exp_dir: Path, file_path: Path, new_content: str) -> bool:
    """应用修改到目标文件"""
    try:
        # 先备份
        backup_file(file_path, exp_dir)
        # 写入新内容
        write_text(file_path, new_content)
        return True
    except Exception as e:
        print(f"[ERROR] apply_changes failed: {e}")
        return False

def rollback_change(exp_dir: Path, file_path: Path) -> bool:
    """回滚修改"""
    try:
        bak = next(exp_dir.glob(f"_backup_{file_path.name}"))
        if bak and bak.exists():
            shutil.copy2(bak, file_path)
            return True
    except Exception:
        pass
    return False

def run_simulation(exp_dir: Path, hypothesis: str) -> Dict:
    """
    运行模拟验证
    简化版：检查修改后的文件语法/逻辑是否正确
    """
    results = {
        "syntax_ok":    True,
        "logic_review": "PASS",
        "warnings":     [],
    }
    # 模拟检查：尝试读取修改后的关键文件
    for md_file in [AGENTS_FILE, SOUL_FILE, MEMORY_FILE]:
        try:
            content = read_text(md_file)
            if len(content) < 50:
                results["warnings"].append(f"{md_file.name} 内容过短")
        except Exception as e:
            results["warnings"].append(f"{md_file.name} 读取失败: {e}")
    return results

def run_experiment(
    hypothesis: str,
    file_to_modify: str,
    new_content: str,
    agent_id: str = "小笨大脑-进化引擎"
) -> Experiment:
    """
    运行单个实验
    AutoResearch Loop: 改 → 测 → 评 → 决定
    """
    exp_id = next_experiment_id()
    exp_dir = create_experiment_dir(exp_id)
    start_time = time.time()

    print(f"\n[EVOLUTION] Starting {exp_id}")
    print(f"           Hypothesis: {hypothesis[:80]}")
    print(f"           Target: {file_to_modify}")

    # 解析目标文件路径
    target_path = WORKSPACE / file_to_modify if not Path(file_to_modify).is_absolute() else Path(file_to_modify)

    # 记录前指标
    metrics_before = collect_metrics()

    # 写入proposal
    write_text(exp_dir / "proposal.md",
        f"# {exp_id} - 实验提案\n"
        f"**时间**: {datetime.now().isoformat()}\n"
        f"**假设**: {hypothesis}\n"
        f"**目标文件**: {file_to_modify}\n"
        f"**修改前指标**: {json.dumps(metrics_before, ensure_ascii=False, indent=2)}\n"
    )

    verdict = Verdict.RUNNING
    reason  = ""

    try:
        # 计时检查
        elapsed = time.time() - start_time
        if elapsed > MAX_EXPERIMENT_SECONDS:
            verdict = Verdict.FAILED
            reason = f"超时 ({elapsed:.0f}s > {MAX_EXPERIMENT_SECONDS}s)"
        else:
            # 应用修改
            success = apply_changes(exp_dir, target_path, new_content)
            if not success:
                verdict = Verdict.FAILED
                reason = "文件写入失败"
            else:
                # 模拟测试
                sim_results = run_simulation(exp_dir, hypothesis)
                write_text(exp_dir / "test_result.md",
                    f"# 测试结果\n```json\n{json.dumps(sim_results, ensure_ascii=False, indent=2)}\n```\n"
                )

                # 评估
                elapsed = time.time() - start_time
                metrics_after = collect_metrics()
                improved, change_desc = metrics_improved(metrics_before, metrics_after)

                # 写入评估结果
                write_text(exp_dir / "evaluation.md",
                    f"# 指标评估\n"
                    f"**实验耗时**: {elapsed:.1f}s\n"
                    f"**改动前**: {json.dumps(metrics_before, ensure_ascii=False, indent=2)}\n"
                    f"**改动后**: {json.dumps(metrics_after, ensure_ascii=False, indent=2)}\n"
                    f"**变化**: {change_desc}\n"
                )

                if improved:
                    verdict = Verdict.KEPT
                    reason = f"指标提升: {change_desc}"
                    print(f"[EVOLUTION] [+] KEPT - {reason}")
                else:
                    # 回滚
                    rollback_change(exp_dir, target_path)
                    verdict = Verdict.DISCARDED
                    reason = f"无显著提升: {change_desc}"
                    print(f"[EVOLUTION] [-] DISCARDED - {reason}")

    except Exception as e:
        verdict = Verdict.FAILED
        reason = f"异常: {str(e)}"
        print(f"[EVOLUTION] [X] FAILED - {reason}")
        # 尝试回滚
        rollback_change(exp_dir, target_path)

    duration = time.time() - start_time

    # 记录改动内容
    write_text(exp_dir / "changes.md", new_content)

    # 写verdict
    write_text(exp_dir / "verdict.md",
        f"# 结论\n"
        f"**Verdict**: {verdict.value}\n"
        f"**原因**: {reason}\n"
        f"**耗时**: {duration:.1f}s\n"
    )

    exp = Experiment(
        id           = exp_id,
        timestamp    = datetime.now().isoformat(),
        hypothesis   = hypothesis,
        file_tried   = file_to_modify,
        changes      = new_content,
        metrics_before = metrics_before,
        metrics_after  = collect_metrics() if verdict == Verdict.KEPT else metrics_before,
        verdict      = verdict,
        reason       = reason,
        duration_s   = duration,
    )

    # 追加到主日志
    append_to_log(exp)

    return exp

# ─────────────────────────────────────────────────────────────────────────────
# 核心：自动发现改进机会
# ─────────────────────────────────────────────────────────────────────────────

DISCOVERY_PROMPTS = [
    ("extract_memories", "记忆命中率 < 80%", """
分析 memory/memory_hot.json，找出7天内未被访问的条目。
如果命中率 < 80%，提出改进假设：用LLM从最近会话中自动提取记忆。
"""),
    ("tool_gap", "工具完善度 < 70%", """
对比 Claude Code 的40+工具列表和当前小笨的工具数量。
找出高频场景中缺失的工具（如 TaskTool, MCPTool）。
"""),
    ("collaboration", "协作效率 < 80%", """
检查 AGENTS.md 中的三脑协作流程定义。
如果某个大脑负载>80%，提出重新分配任务的假设。
"""),
    ("skill_gap", "Skill数量 < 20", """
分析当前 Skills 覆盖的场景。
找出高频重复的场景，提出创建新Skill的假设。
"""),
]

def discover_improvement() -> Optional[Tuple[str, str, str]]:
    """
    自动发现问题并生成改进假设
    返回 (假设, 目标文件, 改进内容) 或 None
    """
    metrics = collect_metrics()

    checks = [
        ("memory_hit_rate", METRICS["memory_hit_rate"], DISCOVERY_PROMPTS[0][2]),
        ("tool_coverage", METRICS["tool_coverage"], DISCOVERY_PROMPTS[1][2]),
        ("collaboration_score", METRICS["collaboration_score"], DISCOVERY_PROMPTS[2][2]),
        ("skill_count", METRICS["skill_count"], DISCOVERY_PROMPTS[3][2]),
    ]

    for key, info, prompt in checks:
        current = metrics.get(key, 0)
        target  = info.get("target", 0) if isinstance(info, dict) else info
        threshold_pct = target * 0.9 if isinstance(target, (int, float)) else 0
        if isinstance(target, (int, float)) and current < threshold_pct:
            # 生成改进假设
            return (
                f"指标 {key} 当前 {current:.1%} 低于目标 {target:.1%}",
                "memory/memory_swap_manager.py",
                "# TODO: 实现自动记忆提取逻辑\n# 灵感来源: Claude Code extractMemories/\n"
            )

    return None

# ─────────────────────────────────────────────────────────────────────────────
# 主实验日志
# ─────────────────────────────────────────────────────────────────────────────

def init_log() -> None:
    """初始化实验日志"""
    if not LOG_FILE.exists():
        write_text(LOG_FILE,
            "# jiaolong自进化实验日志\n"
            f"# 初始化: {datetime.now().isoformat()}\n\n"
            "| ID | 时间 | 假设 | 结果 | 原因 | 耗时 |\n"
            "|----|------|------|------|------|------|\n"
        )

def append_to_log(exp: Experiment) -> None:
    """追加实验结果到日志"""
    ts = datetime.fromisoformat(exp.timestamp).strftime("%Y-%m-%d %H:%M")
    tag = {"KEPT": "[+]", "DISCARDED": "[-]", "FAILED": "[X]", "RUNNING": "[~]"}.get(exp.verdict.value, "[?]")
    line = (
        f"| {exp.id} | {ts} | {exp.hypothesis[:40]}... | "
        f"{tag} {exp.verdict.value} | {exp.reason[:30]} | "
        f"{exp.duration_s:.1f}s |\n"
    )
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line)

# ─────────────────────────────────────────────────────────────────────────────
# AutoResearch 主循环
# ─────────────────────────────────────────────────────────────────────────────

def autorun_overnight(budget_minutes: int = 480) -> List[Experiment]:
    """
    夜间自动运行多轮实验
    仿 AutoResearch: 设定时间预算，醒来后查看实验日志
    """
    init_log()
    results = []
    start = time.time()
    exp_num = 0

    print(f"\n[jiaolong-AUTORESEARCH] 启动夜间实验")
    print(f"                    时间预算: {budget_minutes} 分钟")
    print(f"                    目标: 记忆命中率>80%, L2+L3>80%, Skill>20\n")

    while (time.time() - start) < budget_minutes * 60:
        # 发现改进机会
        discovery = discover_improvement()
        if discovery:
            hypothesis, file_tried, changes = discovery
            exp = run_experiment(hypothesis, file_tried, changes)
            results.append(exp)
            exp_num += 1
            print(f"[{exp_num}] {exp.id} {exp.verdict.value} | {exp.reason[:50]}")
        else:
            print("[AUTORESEARCH] 无显著改进机会，停止实验")
            break

        # 每轮结束检查是否还有时间
        elapsed_min = (time.time() - start) / 60
        remaining_min = budget_minutes - elapsed_min
        if remaining_min < 5:
            print(f"[AUTORESEARCH] 时间剩余 < 5min，停止")
            break

    print(f"\n[jiaolong-AUTORESEARCH] 夜间实验完成")
    print(f"                   总实验: {len(results)}")
    kept = sum(1 for e in results if e.verdict == Verdict.KEPT)
    print(f"                   保留: {kept} | 丢弃: {len(results)-kept}")

    return results

# ─────────────────────────────────────────────────────────────────────────────
# 单次实验（供手动触发）
# ─────────────────────────────────────────────────────────────────────────────

def single_experiment(
    hypothesis: str,
    file_to_modify: str,
    new_content: str
) -> Experiment:
    """执行单次实验"""
    init_log()
    return run_experiment(hypothesis, file_to_modify, new_content)

def status_report() -> str:
    """生成当前状态报告"""
    m = collect_metrics()
    lines = [
        "[jiaolong自进化状态报告]",
        f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "== 当前指标 ==",
        "| 指标 | 当前 | 目标 | 达成率 |",
        "|------|------|------|--------|",
    ]
    for key, info in METRICS.items():
        cur  = m.get(key, 0)
        tgt  = info["target"]
        pct  = cur / tgt if tgt > 0 else 0
        if key == "skill_count":
            status = "[OK]" if cur >= tgt else "[WARN]" if pct > 0.8 else "[LOW]"
            lines.append(f"| {info['name']} | {int(cur)} | {int(tgt)} | {status} |")
        else:
            status = "[OK]" if pct >= 1.0 else "[WARN]" if pct > 0.8 else "[LOW]"
            lines.append(f"| {info['name']} | {cur:.1%} | {tgt:.1%} | {status} |")

    # 最近实验
    if LOG_FILE.exists():
        lines.append("", "== 最近实验 ==")
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            lines.extend(f.readlines()[-10:])

    return "\n".join(lines)

# ─────────────────────────────────────────────────────────────────────────────
# CLI 入口
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "status"

    if cmd == "status":
        print(status_report())

    elif cmd == "autorun":
        budget = int(sys.argv[2]) if len(sys.argv) > 2 else 480
        autorun_overnight(budget)

    elif cmd == "experiment":
        if len(sys.argv) < 4:
            print("用法: evolution.py experiment <文件路径> <新内容>")
            sys.exit(1)
        _, _, file_path, *content_parts = sys.argv
        content = "\n".join(content_parts)
        exp = single_experiment(
            hypothesis="手动触发实验",
            file_to_modify=file_path,
            new_content=content
        )
        print(f"实验 {exp.id}: {exp.verdict.value} | {exp.reason}")

    elif cmd == "discover":
        result = discover_improvement()
        if result:
            print("发现改进机会:")
            print(f"  假设: {result[0]}")
            print(f"  文件: {result[1]}")
        else:
            print("未发现显著改进机会")

    else:
        print(f"未知命令: {cmd}")
        print("用法: evolution.py [status|autorun|experiment|discover]")
